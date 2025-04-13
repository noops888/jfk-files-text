import openpyxl
import csv
import os
import argparse
from tqdm.auto import tqdm # Added for progress bar
from openpyxl.utils import column_index_from_string # To convert letter to index

def extract_links_from_excel(excel_path, csv_path, target_column_letter='A'):
    """
    Extracts filenames (cell text) and hyperlinks from a specified column
    of an Excel sheet and saves them to a CSV file. Assumes Row 1 is the header.
    Reports the specific row number if skipped due to missing data.
    """
    if not os.path.exists(excel_path):
        print(f"Error: Input Excel file not found at {excel_path}")
        return False

    try:
        target_column_index = column_index_from_string(target_column_letter.upper())
        print(f"Processing Excel file: {excel_path} (Target Column: {target_column_letter.upper()})")
    except ValueError:
        print(f"Error: Invalid column letter specified: {target_column_letter}")
        return False

    processed_rows = 0
    skipped_row_reported = False # Flag to report only the first skipped row for clarity
    # skipped_rows = 0 # Now counts rows skipped *after* the header (Removed as we report specific row)

    try:
        print("Loading workbook (this might take a moment)...")
        workbook = openpyxl.load_workbook(excel_path)
        sheet = workbook.active
        print(f"Processing sheet: {sheet.title}")

        if sheet.max_row < 2:
             print("Error: Sheet has less than 2 rows (cannot process header and data).")
             return False

        with open(csv_path, 'w', newline='', encoding='utf-8') as csvfile:
            writer = csv.writer(csvfile)

            writer.writerow(['Filename', 'URL'])
            print("Header written.")

            print(f"Extracting links from rows 2 to {sheet.max_row}...")
            print("--- Debug: First 5 data rows ---")
            for row_index in tqdm(range(2, sheet.max_row + 1), desc="Extracting"):
                # Use the target_column_index
                cell = sheet.cell(row=row_index, column=target_column_index)
                filename_text = cell.value
                url = None

                try:
                    # Attempt to get hyperlink safely
                    if cell.hyperlink and cell.hyperlink.target:
                        url = cell.hyperlink.target.strip()
                except Exception as hyper_err:
                    # Log if there's an error accessing the hyperlink attribute itself
                    if not skipped_row_reported:
                         print(f"\n*** Error accessing hyperlink attribute on row {row_index}: {hyper_err} ***")
                         # We'll treat this as a missing URL for the check below
                         url = None


                if row_index <= 6:
                    print(f"  Row {row_index}: Text='{filename_text}', URL='{url}' (from Col {target_column_letter.upper()})")
                elif row_index == 7:
                     print("  (Debug printing stopped)")

                # Check condition and report specific skipped row if it fails
                if filename_text and url:
                    writer.writerow([filename_text, url])
                    processed_rows += 1
                else:
                    # Report the first skipped row encountered after the header
                    if not skipped_row_reported:
                        reason = []
                        if not filename_text: reason.append("missing text")
                        if not url: reason.append("missing URL (or URL read error)")
                        print(f"\n*** Skipping Excel row {row_index}: Reason(s): {', '.join(reason)}. Text='{filename_text}', URL='{url}' ***")
                        skipped_row_reported = True
                    # We still count total processed rows correctly later if needed
                    # but only report the first skip detail.
                    # skipped_rows += 1 # Removed skipped_rows counter

        print("\n--- Debug End ---")
        print(f"\nExtraction complete.")
        print(f"Output saved to: {csv_path}")
        # Adjust summary message slightly
        print(f"Rows (after header) written to CSV: {processed_rows}")
        if skipped_row_reported:
             print(f"NOTE: At least one data row was skipped (details printed above).")
        else:
             print(f"All data rows (after header) were processed successfully.")

        # Verify row count (CSV Header + Processed Data Rows)
        expected_csv_rows = 1 + processed_rows
        # Optional: Check actual CSV row count if needed, but processed_rows should be accurate
        print(f"Total rows expected in CSV: {expected_csv_rows}")


        return True

    except ImportError:
         print("Error: The 'openpyxl' library is required. Please install it using: pip install openpyxl")
         return False
    except Exception as e:
        print(f"An error occurred during extraction: {e}")
        return False

# --- Main execution ---
if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="Extracts filenames and hyperlinks from a specified column of an Excel file to a CSV.",
        formatter_class=argparse.ArgumentDefaultsHelpFormatter
    )
    parser.add_argument(
        "input_excel",
        help="Path to the input Excel (.xlsx) file."
    )
    parser.add_argument(
        "output_csv",
        help="Path for the output CSV file."
    )
    parser.add_argument(
        "--column",
        default='A',
        help="Column letter (e.g., 'A', 'B') containing the hyperlinked filename."
    )
    args = parser.parse_args()

    if extract_links_from_excel(args.input_excel, args.output_csv, args.column):
        print("CSV generation finished.") # Changed wording slightly
    else:
        print("CSV generation failed.")
        exit(1) 